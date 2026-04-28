"""执行接口测试。"""

from __future__ import annotations

import subprocess
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from typing import Any

import pandas as pd
import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook

from backend.app.api import source_api
from backend.app.api.schemas import DataSource, VariableTag
from backend.app.auth.service import create_access_token, hash_password
from backend.app.database import async_session_factory
from backend.app.loaders.local_reader import load_local_variables
from backend.app.models import User, UserProjectRole
from backend.run import app


TEST_DATA_PATH = Path(__file__).resolve().parent / "data" / "minimal_rules.xlsx"


def _create_composite_test_workbook(target_path: Path) -> Path:
    """创建组合变量测试所需的最小 Excel 文件。"""
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "ID": [1, 2, 3],
                "ItemName": ["Gold", "Wood", "Stone"],
                "Desc": ["黄金+100", "木头+200", "石头+300"],
            }
        ).to_excel(writer, sheet_name="items", index=False)
        pd.DataFrame({"RefID": [1, 2, 9]}).to_excel(
            writer,
            sheet_name="drops",
            index=False,
        )

    return target_path


def _create_dual_composite_test_workbook(target_path: Path) -> Path:
    """创建双组合变量关联比对测试所需的 Excel 文件。"""
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [10001, 10002],
                "INT_ConditionType": [4, 3],
                "INT_RequireRule": [1, 1],
            }
        ).to_excel(writer, sheet_name="left_items", index=False)
        pd.DataFrame(
            {
                "INT_ID": [10001, 10002],
                "INT_ConditionType": [5, 3],
                "INT_RequireRule": [1, 1],
            }
        ).to_excel(writer, sheet_name="right_items", index=False)

    return target_path


def _create_paginated_test_workbook(target_path: Path) -> Path:
    """创建用于执行结果分页测试的 Excel 文件。"""
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        pd.DataFrame({"ID": list(range(1, 11))}).to_excel(
            writer,
            sheet_name="items",
            index=False,
        )
        pd.DataFrame({"RefID": list(range(1, 46))}).to_excel(
            writer,
            sheet_name="drops",
            index=False,
        )

    return target_path


def _create_whitespace_header_workbook(target_path: Path) -> Path:
    """创建带尾部空格 Sheet/列名的 Excel，用于回归原始标识读取链路。"""
    with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": [1001, 1002],
                "STR_ABSwitch  ": ["on", "off"],
                "DESC3": ["开", "关"],
            }
        ).to_excel(writer, sheet_name="Quest  ", index=False)

    return target_path


@pytest.mark.anyio
async def test_execute_engine_returns_three_rule_results() -> None:
    """验证一次请求能同时覆盖空值、重复值和跨表映射缺失。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            },
            {
                "tag": "[drops-ref]",
                "source_id": "src_test",
                "sheet": "drops",
                "column": "RefID",
            },
        ],
        "rules": [
            {
                "rule_type": "not_null",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_type": "unique",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_type": "cross_table_mapping",
                "params": {
                    "dict_tag": "[items-id]",
                    "target_tag": "[drops-ref]",
                },
            },
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["msg"] == "Execution Completed"
    assert payload["meta"]["total_rows_scanned"] > 0
    assert payload["meta"]["failed_sources"] == []

    abnormal_results = payload["data"]["abnormal_results"]
    assert isinstance(abnormal_results, list)
    assert any(
        item["rule_name"] == "not_null" and item["level"] == "error"
        for item in abnormal_results
    )
    assert any(
        item["rule_name"] == "unique" and item["level"] == "warning"
        for item in abnormal_results
    )
    assert any(
        item["rule_name"] == "cross_table_mapping" and item["level"] == "error"
        for item in abnormal_results
    )

    for item in abnormal_results:
        assert set(item) == {
            "level",
            "rule_name",
            "location",
            "row_index",
            "raw_value",
            "message",
        }


@pytest.mark.anyio
async def test_execute_engine_filters_rules_by_selected_ids() -> None:
    """验证 selected_rule_ids 只会执行被勾选的规则。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            },
            {
                "tag": "[drops-ref]",
                "source_id": "src_test",
                "sheet": "drops",
                "column": "RefID",
            },
        ],
        "rules": [
            {
                "rule_id": "rule-not-null",
                "rule_type": "not_null",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_id": "rule-unique",
                "rule_type": "unique",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_id": "rule-cross",
                "rule_type": "cross_table_mapping",
                "params": {
                    "dict_tag": "[items-id]",
                    "target_tag": "[drops-ref]",
                },
            },
        ],
        "selected_rule_ids": ["rule-cross"],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert abnormal_results
    assert all(item["rule_name"] == "cross_table_mapping" for item in abnormal_results)


@pytest.mark.anyio
async def test_execute_engine_returns_400_for_unsupported_rule() -> None:
    """验证未注册规则类型会返回 400。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            }
        ],
        "rules": [{"rule_type": "missing_rule", "params": {}}],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 400
    assert "Unsupported rule_type" in response.json()["detail"]


@pytest.mark.anyio
async def test_execute_engine_returns_400_for_invalid_rule_params() -> None:
    """验证非法规则参数会返回 400。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            }
        ],
        "rules": [
            {
                "rule_type": "not_null",
                "params": {"target_tags": "not-a-list"},
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 400
    assert "params.target_tags" in response.json()["detail"]


@pytest.mark.anyio
async def test_execute_engine_returns_400_for_unknown_source_id() -> None:
    """验证变量引用不存在的 source_id 会返回 400。"""
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            }
        ],
        "variables": [
            {
                "tag": "[bad-source]",
                "source_id": "src_missing",
                "sheet": "items",
                "column": "ID",
            }
        ],
        "rules": [],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 400
    assert "unknown source_id" in response.json()["detail"]


@pytest.mark.anyio
async def test_local_pick_returns_real_selected_path(monkeypatch: pytest.MonkeyPatch) -> None:
    """验证本地文件选择接口会返回真实路径，不会复制文件。"""

    monkeypatch.setattr(
        source_api,
        "_show_local_file_dialog",
        lambda source_type: str(TEST_DATA_PATH) if source_type == "local_excel" else "",
    )

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/local-pick",
            json={"source_type": "local_excel"},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["msg"] == "ok"
    assert payload["data"]["source_type"] == "local_excel"
    assert payload["data"]["selected_path"] == str(TEST_DATA_PATH.resolve())


@pytest.mark.anyio
async def test_local_pick_returns_cancelled_when_user_closes_dialog(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证用户取消选择时接口会返回 cancelled。"""

    monkeypatch.setattr(source_api, "_show_local_file_dialog", lambda _source_type: "")

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/local-pick",
            json={"source_type": "local_excel"},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 204
    assert payload["msg"] == "cancelled"
    assert payload["data"]["selected_path"] == ""


@pytest.mark.anyio
async def test_local_directory_validate_returns_normalized_absolute_directory(
    tmp_path: Path,
) -> None:
    """验证本地目录校验接口会返回规范化后的绝对目录路径。"""

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/local-directory-validate",
            json={"directory_path": f" {tmp_path} "},
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["msg"] == "ok"
    assert payload["data"]["directory_path"] == str(tmp_path.resolve())


def test_show_local_file_dialog_returns_subprocess_stdout(monkeypatch: pytest.MonkeyPatch) -> None:
    """验证子进程方案下，正常返回时取真实路径并去除两端空白。"""

    captured_args: dict[str, Any] = {}

    def fake_run(args: list[str], **kwargs: Any) -> SimpleNamespace:
        captured_args["args"] = args
        captured_args["kwargs"] = kwargs
        return SimpleNamespace(returncode=0, stdout=" C:/tmp/example.xlsx \n", stderr="")

    monkeypatch.setattr(source_api.subprocess, "run", fake_run)

    result = source_api._show_local_file_dialog("local_excel")

    assert result == "C:/tmp/example.xlsx"
    assert captured_args["args"][0] == source_api.sys.executable
    assert captured_args["kwargs"]["timeout"] == source_api._PICKER_SUBPROCESS_TIMEOUT_SECONDS
    assert captured_args["kwargs"]["capture_output"] is True


def test_show_local_file_dialog_returns_empty_when_user_cancelled(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证用户在子进程弹窗中取消时，函数返回空串。"""

    monkeypatch.setattr(
        source_api.subprocess,
        "run",
        lambda *_args, **_kwargs: SimpleNamespace(returncode=0, stdout="", stderr=""),
    )

    assert source_api._show_local_file_dialog("local_csv") == ""


def test_show_local_file_dialog_raises_runtime_error_on_subprocess_timeout(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证子进程超时会被转换为 RuntimeError，由路由层映射成 500。"""

    def raise_timeout(*_args: Any, **_kwargs: Any) -> SimpleNamespace:
        raise subprocess.TimeoutExpired(cmd="python", timeout=1)

    monkeypatch.setattr(source_api.subprocess, "run", raise_timeout)

    with pytest.raises(RuntimeError, match="超时"):
        source_api._show_local_file_dialog("local_excel")


def test_show_local_file_dialog_raises_runtime_error_on_nonzero_exit(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """验证子进程异常退出时，会带上 stderr 抛出 RuntimeError。"""

    monkeypatch.setattr(
        source_api.subprocess,
        "run",
        lambda *_args, **_kwargs: SimpleNamespace(
            returncode=1,
            stdout="",
            stderr="tk init failed",
        ),
    )

    with pytest.raises(RuntimeError, match="tk init failed"):
        source_api._show_local_file_dialog("local_excel")


@pytest.mark.anyio
async def test_source_metadata_returns_excel_sheet_and_column_structure() -> None:
    """验证变量池元数据接口会返回 Excel 的 Sheet 与列结构。"""

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/metadata",
            json={
                "id": "src_test",
                "type": "local_excel",
                "path": str(TEST_DATA_PATH),
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["data"]["source_id"] == "src_test"
    assert payload["data"]["source_type"] == "local_excel"
    assert payload["data"]["sheets"] == [
        {"name": "items", "columns": ["ID", "Name"]},
        {"name": "drops", "columns": ["RefID"]},
    ]


@pytest.mark.anyio
async def test_source_metadata_rejects_csv_for_variable_pool_dropdown(
    tmp_path: Path,
) -> None:
    """验证 CSV 数据源会被明确拦截。"""

    csv_path = tmp_path / "sample.csv"
    csv_path.write_text("id\n1\n", encoding="utf-8")

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/metadata",
            json={
                "id": "src_csv",
                "type": "local_csv",
                "path": str(csv_path),
            },
        )

    assert response.status_code == 400
    assert "变量池下拉提取目前仅支持 Excel 与 SVN" in response.json()["detail"]


@pytest.mark.anyio
async def test_column_preview_returns_top_rows_for_variable_detail() -> None:
    """验证列预览接口会返回变量详情页签所需的前几行数据。"""

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/column-preview",
            json={
                "source": {
                    "id": "src_test",
                    "type": "local_excel",
                    "path": str(TEST_DATA_PATH),
                },
                "sheet": "items",
                "column": "ID",
                "limit": 3,
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["data"]["source_id"] == "src_test"
    assert payload["data"]["sheet"] == "items"
    assert payload["data"]["column"] == "ID"
    assert payload["data"]["preview_limit"] == 3
    assert payload["data"]["total_rows"] == 5
    assert payload["data"]["preview_rows"] == [
        {"row_index": 2, "value": 1},
        {"row_index": 3, "value": 2},
        {"row_index": 4, "value": 2},
    ]


@pytest.mark.anyio
async def test_column_preview_without_limit_returns_full_column_for_detail_dialog() -> None:
    """验证详情弹窗在不传 limit 时会返回当前列的完整预览。"""

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/column-preview",
            json={
                "source": {
                    "id": "src_test",
                    "type": "local_excel",
                    "path": str(TEST_DATA_PATH),
                },
                "sheet": "items",
                "column": "ID",
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["data"]["source_path"] == str(TEST_DATA_PATH)
    assert payload["data"]["total_rows"] == 5
    assert payload["data"]["loaded_rows"] == 5
    assert payload["data"]["loaded_all_rows"] is True
    assert payload["data"]["preview_limit"] == 5
    assert payload["data"]["preview_rows"] == [
        {"row_index": 2, "value": 1},
        {"row_index": 3, "value": 2},
        {"row_index": 4, "value": 2},
        {"row_index": 5, "value": None},
        {"row_index": 6, "value": "   "},
    ]


@pytest.mark.anyio
async def test_composite_preview_returns_json_mapping_for_same_sheet(
    tmp_path: Path,
) -> None:
    """验证组合变量预览接口会返回 key->object 的完整 JSON 映射。"""

    workbook_path = _create_composite_test_workbook(tmp_path / "composite_preview.xlsx")

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/composite-preview",
            json={
                "source": {
                    "id": "src_combo",
                    "type": "local_excel",
                    "path": str(workbook_path),
                },
                "sheet": "items",
                "columns": ["ID", "ItemName", "Desc"],
                "key_column": "ID",
            },
        )

    assert response.status_code == 200
    payload = response.json()
    assert payload["code"] == 200
    assert payload["data"]["variable_kind"] == "composite"
    assert payload["data"]["sheet"] == "items"
    assert payload["data"]["columns"] == ["ID", "ItemName", "Desc"]
    assert payload["data"]["key_column"] == "ID"
    assert payload["data"]["has_duplicate_keys"] is False
    assert payload["data"]["duplicate_keys_preview"] == []
    assert payload["data"]["total_rows"] == 3
    assert payload["data"]["loaded_rows"] == 3
    assert payload["data"]["mapping"] == {
        "1": {"ItemName": "Gold", "Desc": "黄金+100"},
        "2": {"ItemName": "Wood", "Desc": "木头+200"},
        "3": {"ItemName": "Stone", "Desc": "石头+300"},
    }


@pytest.mark.anyio
async def test_composite_preview_supports_appending_index_to_duplicate_keys(
    tmp_path: Path,
) -> None:
    """验证组合变量预览可将重复 key 生成为原值_序号。"""
    workbook_path = tmp_path / "composite_duplicate_keys.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "INT_ID": ["A", "A", "B"],
                "INT_Group": [1, 2, 3],
                "INT_Faction": [0, 1, 1],
            }
        ).to_excel(writer, sheet_name="items", index=False)

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/composite-preview",
            json={
                "source": {
                    "id": "src_combo",
                    "type": "local_excel",
                    "path": str(workbook_path),
                },
                "sheet": "items",
                "columns": ["INT_ID", "INT_Group", "INT_Faction"],
                "key_column": "INT_ID",
                "append_index_to_key": True,
            },
        )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["has_duplicate_keys"] is True
    assert payload["duplicate_keys_preview"] == ["A"]
    assert payload["mapping"] == {
        "A_0": {"INT_Group": 1, "INT_Faction": 0},
        "A_1": {"INT_Group": 2, "INT_Faction": 1},
        "B_2": {"INT_Group": 3, "INT_Faction": 1},
    }


@pytest.mark.anyio
async def test_composite_preview_reports_duplicate_keys_without_append_mode(
    tmp_path: Path,
) -> None:
    """验证重复 key 且未开启追加序号时，预览仍返回重复标记供前端提示。"""
    workbook_path = tmp_path / "composite_duplicate_keys_without_append.xlsx"
    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "STR_ParamType": ["list", "list", "other"],
                "INT_Group": [1, 2, 3],
                "INT_Faction": [0, 1, 1],
            }
        ).to_excel(writer, sheet_name="switch", index=False)

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post(
            "/api/v1/sources/composite-preview",
            json={
                "source": {
                    "id": "src_combo",
                    "type": "local_excel",
                    "path": str(workbook_path),
                },
                "sheet": "switch",
                "columns": ["STR_ParamType", "INT_Group", "INT_Faction"],
                "key_column": "STR_ParamType",
                "append_index_to_key": False,
            },
        )

    assert response.status_code == 200
    payload = response.json()["data"]
    assert payload["has_duplicate_keys"] is True
    assert payload["duplicate_keys_preview"] == ["list"]
    assert payload["mapping"] == {}
    assert payload["loaded_rows"] == 0


@pytest.mark.anyio
async def test_preview_endpoints_preserve_raw_sheet_and_column_names(
    tmp_path: Path,
) -> None:
    """验证预览接口读取 Excel 时保留原始 Sheet 名和列名，不裁掉尾部空格。"""
    workbook_path = _create_whitespace_header_workbook(tmp_path / "whitespace_headers.xlsx")

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        column_response = await client.post(
            "/api/v1/sources/column-preview",
            json={
                "source": {
                    "id": "src_space",
                    "type": "local_excel",
                    "path": str(workbook_path),
                },
                "sheet": "Quest  ",
                "column": "STR_ABSwitch  ",
            },
        )
        composite_response = await client.post(
            "/api/v1/sources/composite-preview",
            json={
                "source": {
                    "id": "src_space",
                    "type": "local_excel",
                    "path": str(workbook_path),
                },
                "sheet": "Quest  ",
                "columns": ["INT_ID", "STR_ABSwitch  ", "DESC3"],
                "key_column": "INT_ID",
            },
        )

    assert column_response.status_code == 200
    column_payload = column_response.json()["data"]
    assert column_payload["sheet"] == "Quest  "
    assert column_payload["column"] == "STR_ABSwitch  "
    assert column_payload["preview_rows"] == [
        {"row_index": 2, "value": "on"},
        {"row_index": 3, "value": "off"},
    ]

    assert composite_response.status_code == 200
    composite_payload = composite_response.json()["data"]
    assert composite_payload["sheet"] == "Quest  "
    assert composite_payload["columns"] == ["INT_ID", "STR_ABSwitch  ", "DESC3"]
    assert composite_payload["mapping"] == {
        "1001": {"STR_ABSwitch  ": "on", "DESC3": "开"},
        "1002": {"STR_ABSwitch  ": "off", "DESC3": "关"},
    }


@pytest.mark.anyio
async def test_execute_engine_supports_trimmed_sheet_and_column_identifiers(
    tmp_path: Path,
) -> None:
    """验证执行链路可兼容已被 trim 的 Sheet/列配置，并正确解析真实表头。"""
    workbook_path = _create_whitespace_header_workbook(tmp_path / "trimmed_execute.xlsx")
    payload = {
        "sources": [
            {
                "id": "src_space",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[quest-switch]",
                "source_id": "src_space",
                "sheet": "Quest",
                "column": "STR_ABSwitch",
            }
        ],
        "rules": [
            {
                "rule_type": "not_null",
                "params": {"target_tags": ["[quest-switch]"]},
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    response_payload = response.json()
    assert response_payload["meta"]["failed_sources"] == []
    assert response_payload["meta"]["total_rows_scanned"] == 2
    assert response_payload["data"]["abnormal_results"] == []


def test_load_local_variables_preserves_raw_excel_identifiers(
    tmp_path: Path,
) -> None:
    """验证执行链路加载变量时也保留原始 Sheet 名和列名。"""
    workbook_path = _create_whitespace_header_workbook(tmp_path / "whitespace_execute.xlsx")

    loaded_variables = load_local_variables(
        [
            DataSource(
                id="src_space",
                type="local_excel",
                path=str(workbook_path),
            )
        ],
        [
            VariableTag(
                tag="[quest-switch-single]",
                source_id="src_space",
                sheet="Quest  ",
                variable_kind="single",
                column="STR_ABSwitch  ",
                expected_type="str",
            ),
            VariableTag(
                tag="[quest-switch-composite]",
                source_id="src_space",
                sheet="Quest  ",
                variable_kind="composite",
                columns=["INT_ID", "STR_ABSwitch  ", "DESC3"],
                key_column="INT_ID",
                expected_type="json",
            ),
        ],
    )

    single_frame = loaded_variables["[quest-switch-single]"]
    assert single_frame.columns.tolist() == ["STR_ABSwitch  ", "_row_index"]
    assert single_frame["STR_ABSwitch  "].tolist() == ["on", "off"]

    composite_frame = loaded_variables["[quest-switch-composite]"]
    assert composite_frame.columns.tolist() == ["__key__", "STR_ABSwitch  ", "DESC3", "_row_index"]
    assert composite_frame["__key__"].tolist() == ["1001", "1002"]


@pytest.mark.anyio
async def test_execute_engine_accepts_composite_variable_without_breaking_rules(
    tmp_path: Path,
) -> None:
    """验证 TaskTree 中包含组合变量时，现有三类规则仍可对单变量正常执行。"""

    workbook_path = _create_composite_test_workbook(tmp_path / "composite_execute.xlsx")

    payload = {
        "sources": [
            {
                "id": "src_combo",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_combo",
                "sheet": "items",
                "variable_kind": "single",
                "column": "ID",
            },
            {
                "tag": "[drops-ref]",
                "source_id": "src_combo",
                "sheet": "drops",
                "variable_kind": "single",
                "column": "RefID",
            },
            {
                "tag": "[items-json]",
                "source_id": "src_combo",
                "sheet": "items",
                "variable_kind": "composite",
                "columns": ["ID", "ItemName", "Desc"],
                "key_column": "ID",
                "expected_type": "json",
            },
        ],
        "rules": [
            {
                "rule_type": "not_null",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_type": "unique",
                "params": {"target_tags": ["[items-id]"]},
            },
            {
                "rule_type": "cross_table_mapping",
                "params": {
                    "dict_tag": "[items-id]",
                    "target_tag": "[drops-ref]",
                },
            },
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    response_payload = response.json()
    assert response_payload["msg"] == "Execution Completed"
    assert response_payload["meta"]["failed_sources"] == []
    assert response_payload["meta"]["total_rows_scanned"] > 0
    assert any(
        item["rule_name"] == "cross_table_mapping" and item["raw_value"] == 9
        for item in response_payload["data"]["abnormal_results"]
    )


@pytest.mark.anyio
async def test_execute_engine_supports_dual_composite_compare(tmp_path: Path) -> None:
    """验证个人校验执行接口支持双组合变量按 Key 关联后比较字段值。"""
    workbook_path = _create_dual_composite_test_workbook(tmp_path / "dual_composite_rules.xlsx")

    payload = {
        "sources": [
            {
                "id": "src_dual",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[left-json]",
                "source_id": "src_dual",
                "sheet": "left_items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_ConditionType", "INT_RequireRule"],
                "key_column": "INT_ID",
                "expected_type": "json",
            },
            {
                "tag": "[right-json]",
                "source_id": "src_dual",
                "sheet": "right_items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_ConditionType", "INT_RequireRule"],
                "key_column": "INT_ID",
                "expected_type": "json",
            },
        ],
        "rules": [
            {
                "rule_id": "rule-dual",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[left-json]",
                    "reference_tag": "[right-json]",
                    "key_check_mode": "baseline_only",
                    "rule_name": "双组合变量比对",
                    "comparisons": [
                        {
                            "comparison_id": "compare-condition-type",
                            "left_field": "INT_ConditionType",
                            "operator": "eq",
                            "right_field": "INT_ConditionType",
                        },
                        {
                            "comparison_id": "compare-require-rule",
                            "left_field": "INT_RequireRule",
                            "operator": "eq",
                            "right_field": "INT_RequireRule",
                        },
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    abnormal_results = response.json()["data"]["abnormal_results"]
    assert len(abnormal_results) == 1
    assert abnormal_results[0]["rule_name"] == "双组合变量比对"
    assert "Key 10001" in abnormal_results[0]["message"]


@pytest.mark.anyio
async def test_execute_engine_uses_appended_composite_keys_during_runtime(tmp_path: Path) -> None:
    """验证执行链路中的组合变量 __key__ 与预览使用相同的原值_序号口径。"""
    workbook_path = _create_dual_composite_test_workbook(tmp_path / "dual_composite_suffix_keys.xlsx")

    payload = {
        "sources": [
            {
                "id": "src_dual",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[left-json]",
                "source_id": "src_dual",
                "sheet": "left_items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_ConditionType", "INT_RequireRule"],
                "key_column": "INT_ID",
                "append_index_to_key": True,
                "expected_type": "json",
            },
            {
                "tag": "[right-json]",
                "source_id": "src_dual",
                "sheet": "right_items",
                "variable_kind": "composite",
                "columns": ["INT_ID", "INT_ConditionType", "INT_RequireRule"],
                "key_column": "INT_ID",
                "append_index_to_key": True,
                "expected_type": "json",
            },
        ],
        "rules": [
            {
                "rule_id": "rule-dual",
                "rule_type": "dual_composite_compare",
                "params": {
                    "target_tag": "[left-json]",
                    "reference_tag": "[right-json]",
                    "key_check_mode": "baseline_only",
                    "rule_name": "双组合变量比对",
                    "comparisons": [
                        {
                            "comparison_id": "compare-suffixed-key",
                            "left_field": "__key__",
                            "operator": "eq",
                            "right_field": "__key__",
                        }
                    ],
                },
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200
    assert response.json()["data"]["abnormal_results"] == []


@pytest.mark.anyio
async def test_execute_engine_persists_latest_result_and_supports_server_side_pagination(
    tmp_path: Path,
    auth_headers: dict[str, str],
    test_project_id: int,
) -> None:
    """验证个人校验执行后会返回第一页结果，并支持按 result_id 翻页。"""
    workbook_path = _create_paginated_test_workbook(tmp_path / "paged_execute.xlsx")
    payload = {
        "sources": [
            {
                "id": "src_test",
                "type": "local_excel",
                "path": str(workbook_path),
            }
        ],
        "variables": [
            {
                "tag": "[items-id]",
                "source_id": "src_test",
                "sheet": "items",
                "column": "ID",
            },
            {
                "tag": "[drops-ref]",
                "source_id": "src_test",
                "sheet": "drops",
                "column": "RefID",
            },
        ],
        "rules": [
            {
                "rule_id": "rule-cross",
                "rule_type": "cross_table_mapping",
                "params": {
                    "dict_tag": "[items-id]",
                    "target_tag": "[drops-ref]",
                },
            }
        ],
        "selected_rule_ids": ["rule-cross"],
        "page": 1,
        "size": 20,
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
        headers=auth_headers,
    ) as client:
        execute_response = await client.post("/api/v1/engine/execute", json=payload)

        assert execute_response.status_code == 200
        execute_payload = execute_response.json()
        assert execute_payload["msg"] == "Execution Completed"
        assert execute_payload["meta"]["result_id"] > 0
        assert execute_payload["data"]["page"] == 1
        assert execute_payload["data"]["size"] == 20
        assert execute_payload["data"]["total"] == 35
        assert len(execute_payload["data"]["list"]) == 20
        assert execute_payload["data"]["abnormal_results"] == execute_payload["data"]["list"]

        result_id = execute_payload["meta"]["result_id"]
        page_two_response = await client.get(
            f"/api/v1/engine/results/{result_id}",
            params={"page": 2, "size": 20},
        )
        export_response = await client.get(f"/api/v1/engine/results/{result_id}/export")

    assert page_two_response.status_code == 200
    page_two_payload = page_two_response.json()
    assert page_two_payload["meta"]["result_id"] == result_id
    assert page_two_payload["data"]["page"] == 2
    assert page_two_payload["data"]["size"] == 20
    assert page_two_payload["data"]["total"] == 35
    assert len(page_two_payload["data"]["list"]) == 15
    assert page_two_payload["data"]["list"][0]["raw_value"] == 31

    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(export_response.content))
    assert workbook.sheetnames == ["统计摘要", "异常明细"]
    assert workbook["统计摘要"]["B2"].value == "个人校验"
    assert workbook["异常明细"].max_row == 36
    assert workbook["异常明细"]["A1"].value == "级别"

    async with async_session_factory() as session:
        other_user = User(
            username="other-workbench-user",
            hashed_password=hash_password("testpass"),
            primary_project_id=test_project_id,
        )
        session.add(other_user)
        await session.flush()
        session.add(
            UserProjectRole(
                user_id=other_user.id,
                project_id=test_project_id,
                role="user",
            )
        )
        await session.commit()
        other_token = create_access_token(other_user.id, project_id=test_project_id)

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
        headers={"Authorization": f"Bearer {other_token}"},
    ) as client:
        forbidden_export_response = await client.get(
            f"/api/v1/engine/results/{result_id}/export"
        )

    assert forbidden_export_response.status_code == 404


@pytest.mark.anyio
async def test_execute_engine_runs_svn_remote_source(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """SVN 远端 URL 数据源走 prepare_remote_svn_source 落到本地 fixture，
    应当与等价的 local_excel 配置产出相同的异常结果。"""
    from backend.app.loaders import svn_cache

    fixture_xlsx = tmp_path / "remote_quests.xlsx"
    pd.DataFrame({"ID": [1, 2, 3], "Name": ["A", "", "C"]}).to_excel(
        fixture_xlsx,
        sheet_name="items",
        index=False,
    )

    def _fake_prepare(source, *, user_scope=None, force_refresh=False):
        return fixture_xlsx

    monkeypatch.setattr(svn_cache, "prepare_remote_svn_source", _fake_prepare)

    payload = {
        "sources": [
            {
                "id": "src_remote",
                "type": "svn",
                "pathOrUrl": "https://samosvn/data/project/samo/GameDatas/datas_qa88/remote_quests.xlsx",
            }
        ],
        "variables": [
            {"tag": "[items-name]", "source_id": "src_remote", "sheet": "items", "column": "Name"}
        ],
        "rules": [
            {
                "rule_id": "rule_not_null_name",
                "rule_type": "not_null",
                "params": {"target_tags": ["[items-name]"]},
            }
        ],
    }

    async with AsyncClient(
        transport=ASGITransport(app=app),
        base_url="http://testserver",
    ) as client:
        response = await client.post("/api/v1/engine/execute", json=payload)

    assert response.status_code == 200, response.text
    body = response.json()
    abnormal_results = body["data"]["abnormal_results"]
    assert any(item["rule_name"] for item in abnormal_results)
    # 至少能命中一行空 Name 异常
    assert len(abnormal_results) >= 1
    failed_sources = body["meta"]["failed_sources"]
    assert "src_remote" not in failed_sources
