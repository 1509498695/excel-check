"""执行结果 Excel 导出工具。"""

from __future__ import annotations

import json
from datetime import datetime
from io import BytesIO
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


RESULT_EXPORT_MIME_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


def build_execution_result_workbook(
    payload: dict[str, Any],
    *,
    scope_label: str,
) -> BytesIO:
    """把完整执行结果转换成包含摘要和明细的 Excel 工作簿。"""
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "统计摘要"
    detail_sheet = workbook.create_sheet("异常明细")

    _write_summary_sheet(summary_sheet, payload, scope_label)
    _write_detail_sheet(detail_sheet, payload.get("list", []))

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _write_summary_sheet(sheet: Any, payload: dict[str, Any], scope_label: str) -> None:
    failed_sources = payload.get("failed_sources") or []
    rows = [
        ("校验类型", scope_label),
        ("结果 ID", payload.get("result_id", "")),
        ("执行时间", _format_cell_value(payload.get("created_at"))),
        ("扫描总行数", payload.get("total_rows_scanned", 0)),
        ("失败数据源", "、".join(str(item) for item in failed_sources)),
        ("异常结果数", payload.get("total", 0)),
        ("执行耗时(ms)", payload.get("execution_time_ms", 0)),
    ]

    sheet.append(["指标", "值"])
    for row in rows:
        sheet.append(list(row))

    _style_header(sheet, 1)
    _auto_fit_columns(sheet)


def _write_detail_sheet(sheet: Any, abnormal_results: list[dict[str, Any]]) -> None:
    headers = ["级别", "规则名称", "定位", "行号", "原始值", "说明"]
    sheet.append(headers)
    for item in abnormal_results:
        sheet.append(
            [
                _format_cell_value(item.get("level")),
                _format_cell_value(item.get("rule_name")),
                _format_cell_value(item.get("location")),
                item.get("row_index", ""),
                _format_cell_value(item.get("raw_value")),
                _format_cell_value(item.get("message")),
            ]
        )

    sheet.freeze_panes = "A2"
    _style_header(sheet, 1)
    _auto_fit_columns(sheet)


def _style_header(sheet: Any, row_index: int) -> None:
    fill = PatternFill("solid", fgColor="EAF2FF")
    font = Font(bold=True, color="1F2A44")
    for cell in sheet[row_index]:
        cell.fill = fill
        cell.font = font


def _auto_fit_columns(sheet: Any) -> None:
    for column_cells in sheet.columns:
        max_length = 0
        column_index = column_cells[0].column
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or "")))
        width = min(max(max_length + 2, 12), 60)
        sheet.column_dimensions[get_column_letter(column_index)].width = width


def _format_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)
